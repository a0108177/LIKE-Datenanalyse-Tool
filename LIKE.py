# %%
# Bibliotheken importieren: pandas für Datenverarbeitung, OS/Datum für Dateinamen
import pandas as pd
import os
from datetime import datetime

# # %% [markdown]
# # # Import

# # %%
# # Daten für ausgewähltes Land einlesen und Parameter setzen
# # Region und number of enrollments festlegen
# country = "CZ"
# enrolled = 110

# # CSVs einlesen (UTF-8) --> Pfad anpassen
# df_large_table = pd.read_csv(
#     f"Countries/{country}/LINDE MH _ KION - LARGE TABLE (ADAPTIVE) (V2).csv",
#     sep=";",
#     encoding="utf-8",
# )
# df_mcp = pd.read_csv(
#     f"Countries/{country}/Metacognition Progress.csv", sep=";", encoding="utf-8"
# )
# df_self_assessment = pd.read_csv(
#     f"Countries/{country}/COMPLETED LEARNERS WITH SELF-ASSESSMENT BY MODULE.csv",
#     sep=";",
#     encoding="utf-8",
# )
# df_mdlo = pd.read_csv(
#     f"Countries/{country}/THE MOST DIFFICULT LEARNING OBJECTIVES.csv",
#     sep=";",
#     encoding="utf-8",
# )

# # %% [markdown]
# # # Transform Large Table

# # %%
# # Wir wollen nur die Learner haben die alle Module abgeschlossen haben
def like(df_large_table, df_mcp, df_self_assessment, df_mdlo):
    # "Completion Status" ist "Completed"
    df_large_table = df_large_table[
        df_large_table["Completion Status"] == "COMPLETED"
    ].copy()
    # Namen der Module
    modules_names = set(df_large_table["Module"].unique())
    # Liste welcher Learner hat welche Module abgeschlossen
    modules_per_learner = df_large_table.groupby("Learner")["Module"].apply(set)
    # Liste mit Namen der Learner die alle Module abgeschlossen haben
    completed_learners = modules_per_learner[modules_per_learner ==
                                            modules_names].index
    df_large_table = df_large_table[df_large_table["Learner"].isin(
        completed_learners)]

    # Large Table enthält Spalten die als Strings interpretiert werden -> Typen Umwandlung
    columns_with_percent_values = [
        "Accuracy",
        "Unconscious Incompetent",
        "Conscious Incompetent",
        "Unconscious Competent",
        "Conscious Competent",
    ]

    for col in columns_with_percent_values:
        df_large_table[col] = (
            df_large_table[col].astype(str).str.replace(
                "%", "", regex=False).str.strip()
        )
        df_large_table[col] = pd.to_numeric(df_large_table[col], errors="coerce")
    df_large_table

    # %% [markdown]
    # #

    # %% [markdown]
    # # Amount Of Learners

    # %%
    amount_of_learners = len(completed_learners)

    df_amount_of_learners = pd.DataFrame(
        {
            "Description": ["Amount Of Participants Who Completed All Modules"],
            "Value": [amount_of_learners],
        }
    )
    df_amount_of_learners

    # %% [markdown]
    # # Time To Complete

    # %%
    # Hilfsfunktion: timedelta in hh mm umwandeln
    def td_to_hm(td):
        total_minutes = int(td.total_seconds() // 60)
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h}h {m}m"


    df_large_table["Sum Time Spent"] = pd.to_timedelta(
        df_large_table["Sum Time Spent"])

    df_time = (
        df_large_table.groupby(["Class Description", "Learner"])["Sum Time Spent"]
        .sum()
        .reset_index()
    )

    results = []
    for class_desc, group in df_time.groupby("Class Description"):
        mean_time = td_to_hm(group["Sum Time Spent"].mean())
        min_time = td_to_hm(group["Sum Time Spent"].min())
        max_time = td_to_hm(group["Sum Time Spent"].max())
        sum_time = td_to_hm(group["Sum Time Spent"].sum())
        results.append(
            {
                "Class Description": class_desc,
                "Average Time Needed To Complete All Modules": mean_time,
                "Min. Time Needed To Complete All Modules": min_time,
                "Max. Time Needed To Complete All Modules": max_time,
                "Total Time Needed To Complete All Modules": sum_time,
            }
        )

    df_time = pd.DataFrame(results)
    df_time

    # %% [markdown]
    # # Average Accuracy (All Modules)

    # %%
    df_large_table["Accuracy"] = pd.to_numeric(
        df_large_table["Accuracy"], errors="coerce")

    # calculate mean per class
    accuracy_by_class = (
        df_large_table.groupby("Class Description")["Accuracy"]
        .mean()
        .round(2)
        .reset_index()
    )

    df_avg_acc = pd.DataFrame(
        {
            "Class Description": [
                str(row["Class Description"]) for idx, row in accuracy_by_class.iterrows()
            ],
            "Average Accuracy (all modules)": accuracy_by_class["Accuracy"],
        }
    )
    df_avg_acc

    # %% [markdown]
    # # Metacognition Progress

    # %%
    # teils auch negativ Werte dabei (= reduction)
    def convert(value):
        return int(abs(value) * 100)


    # Initial Consciousness
    # = Consciousness
    cc = convert(df_mcp["Initial Conscious Competence"].iloc[0])
    uc = convert(df_mcp["Initial Unconscious Competence"].iloc[0])
    ci = convert(df_mcp["Initial Conscious Incompetence"].iloc[0])
    ui = convert(df_mcp["Initial Unconscious Incompetence"].iloc[0])

    # Improvement Conscious Competence
    # = Knowledge increase (Metacognition Progress)
    icc = df_mcp["Improvement Conscious Competence"].iloc[0]
    icc = convert(icc)

    # Improvement Unconscious Incompetence
    # = Reduction Unconscious Incompetence
    iui = df_mcp["Improvement Unconscious Incompetence"].iloc[0]
    iui = convert(iui)

    metacognition_progress_labels = [
        "Initial Unconscious Incompetence (UI)",
        "Initial Conscious Incompetence (CI)",
        "Initial Unconscious Competence (UC)",
        "Initial Conscious Competence (CC)",
        "Knowledge increase (Metacognition Progress)",
        "Reduction Unconscious Incompetence",
    ]
    metacognition_progress_values = [ui, ci, uc, cc, icc, iui]

    df_mcp = pd.DataFrame(
        {
            "Description": metacognition_progress_labels,
            "Value": metacognition_progress_values,
        }
    )
    df_mcp

    # %% [markdown]
    # # Accuracy Of Self-Assessment

    # %%
    order_acc = [
        "firm knowledge",
        "competent; training voluntary - no immediate need",
        "profits from re-training / webinar",
        "hands-on classroom training needed",
    ]

    acc_of_self_assessment = []

    modules = sorted(df_large_table["Module"].unique())

    for module in modules:
        subset = df_large_table[df_large_table["Module"] == module]
        percentages_acc = (
            subset["Accuracy-Classes"].value_counts(normalize=True) * 100
        ).round(2)
        row = {"Module": module}
        for value in order_acc:
            row[value] = percentages_acc.get(value, 0)
        acc_of_self_assessment.append(row)

    df_acc_of_self_assessment = pd.DataFrame(acc_of_self_assessment)

    # Summary Order ->
    df_acc_of_self_assessment[">69%"] = round(
        df_acc_of_self_assessment["firm knowledge"]
        + df_acc_of_self_assessment["competent; training voluntary - no immediate need"],
        2,
    )
    df_acc_of_self_assessment["<=69% - >50%"] = round(
        df_acc_of_self_assessment["profits from re-training / webinar"], 2
    )
    df_acc_of_self_assessment["<= 50%"] = round(
        df_acc_of_self_assessment["hands-on classroom training needed"], 2
    )
    df_acc_of_self_assessment

    # %% [markdown]
    # # Self-evaluation

    # %%
    # Selbsteinschätzung je Modul: Häufigkeiten (in %) berechnen und zusammenfassen
    order_sa = ["Novice", "Advanced beginner", "Competent", "Proficient", "Expert"]

    self_assessment = []

    modules = sorted(df_large_table["Module"].unique())

    for module in modules:
        subset = df_self_assessment[df_self_assessment["Module"] == module]
        # Prozentuale Verteilung je Selbsteinschätzungsstufe
        percentages_sa = (
            subset["Self Assessment"].value_counts(normalize=True) * 100
        ).round(2)
        row = {"Module": module}
        for value in order_sa:
            row[value] = percentages_sa.get(value, 0)
        self_assessment.append(row)

    df_self_assessment = pd.DataFrame(self_assessment)

    # Zusammenfassung: Professional (Proficient+Expert), Competent, Beginner (Novice+Advanced beginner)
    df_self_assessment["Professional"] = round(
        df_self_assessment["Proficient"] + df_self_assessment["Expert"], 2
    )
    df_self_assessment["Competent 2"] = round(df_self_assessment["Competent"], 2)
    df_self_assessment["Beginner"] = round(
        df_self_assessment["Novice"] + df_self_assessment["Advanced beginner"], 2
    )
    df_self_assessment

    # %% [markdown]
    # # Competence Level

    # %%
    # Kompetenzniveau je Modul aggregieren und in Summen-Kategorien zusammenfassen
    competence_columns = [
        "Unconscious Incompetent",
        "Conscious Incompetent",
        "Unconscious Competent",
        "Conscious Competent",
    ]

    # Mittelwerte der Kompetenzstufen je Modul berechnen
    df_competence_by_module = (
        df_large_table.groupby("Module")[
            competence_columns].mean().round(2).reset_index()
        )

    # Zusammenfassung: Incompetent = UI + CI, Competent = UC + CC
    df_competence_by_module["Incompetent"] = round(
        df_competence_by_module["Unconscious Incompetent"]
        + df_competence_by_module["Conscious Incompetent"],
        2,
    )
    df_competence_by_module["Competent"] = round(
        df_competence_by_module["Unconscious Competent"]
        + df_competence_by_module["Conscious Competent"],
        2,
    )
    df_competence_by_module

    # %% [markdown]
    # # Top 5 Most Difficult Learning Objectives

    # %%
    df_mdlo = df_mdlo.head(5)
    # wir wollen nach UI sortieren -> Grudlage zur Bewertung der Learning Objectives
    df_mdlo = df_mdlo.sort_values(by="Unconsciously Incompetent", ascending=False)
    df_mdlo = df_mdlo.drop(columns=["Open in Curator"])
    df_mdlo

    # %% [markdown]
    # # Format

    # %%
    # Für eine einzelne Spalte:
    df_avg_acc["Average Accuracy (all modules)"] = (
        df_avg_acc["Average Accuracy (all modules)"].astype(
            str).str.replace(".", ",") + "%"
    )

    df_mcp["Value"] = df_mcp["Value"].astype(str).str.replace(".", ",") + "%"

    # Für alle relevanten Spalten (außer "Module") im DataFrame:
    for col in df_acc_of_self_assessment.columns:
        if col == "Module":
            continue
        df_acc_of_self_assessment[col] = (
            df_acc_of_self_assessment[col].astype(str).str.replace(".", ",") + "%"
        )

    for col in df_self_assessment.columns:
        if col == "Module":
            continue
        df_self_assessment[col] = (
            df_self_assessment[col].astype(str).str.replace(".", ",") + "%"
        )

    for col in df_competence_by_module.columns:
        if col == "Module":
            continue
        df_competence_by_module[col] = (
            df_competence_by_module[col].astype(str).str.replace(".", ",") + "%"
        )

    # %% [markdown]
    # # Export In .xlsx

    # %%
    # Export der Ergebnisse in eine Excel-Datei mit heutigem Datum
    from openpyxl import load_workbook
    today_str = datetime.now().strftime("%Y%m%d")
    filename = f"{today_str}_Like_Auswertung.xlsx"
    # Schreibe mehrere DataFrames auf separate Tabellenblätter
    with pd.ExcelWriter(filename) as writer:
        # Teilnehmer (abgeschlossen)
        df_amount_of_learners.to_excel(
            writer, sheet_name="Participants Completed", index=False
        )
        # Benötigte Zeit je Einheit
        df_time.to_excel(writer, sheet_name="Time Needed", index=False)
        # Durchschnittliche Genauigkeit
        df_avg_acc.to_excel(writer, sheet_name="Avg Accuracy", index=False)
        # Metakognition-Fortschritt
        df_mcp.to_excel(writer, sheet_name="Metacognition Progress", index=False)
        # Genauigkeit pro Modul (Selbsteinschätzung)
        df_acc_of_self_assessment.to_excel(
            writer, sheet_name="Accuracy Per Module", index=False
        )
        # Selbsteinschätzung pro Modul
        df_self_assessment.to_excel(
            writer, sheet_name="Self Assessment Per Module", index=False
        )
        # Kompetenzniveau pro Modul
        df_competence_by_module.to_excel(
            writer, sheet_name="Competence Level Per Module", index=False
        )
        # Schwierigste Lernziele
        df_mdlo.to_excel(
            writer, sheet_name="5 Most Dfficult Objectives", index=False)

    # Spaltenbreiten in allen Tabellenblättern automatisch anpassen
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max(max_length + 2, 10)
    wb.save(filename)